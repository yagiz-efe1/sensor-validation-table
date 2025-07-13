
# Tkinter ile kullanıcıdan veri alarak sensör tablosu oluşturur ve Excel dosyası olarak kaydeder.

import tkinter as tk
from tkinter import simpledialog, messagebox
import pandas as pd
import itertools
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os

# Excel dosyasını biçimlendirme fonksiyonu
def format_excel(filepath):
    wb = load_workbook(filepath)
    ws = wb.active
    ws.insert_rows(1)  # Grup başlıkları için en üste boş bir satır ekle

    # Renk ve hücre stilleri
    header_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # Sarı başlık rengi
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")     # 1 için kırmızı
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")   # 0 için yeşil
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))  # İnce kenarlık

    total_cols = ws.max_column
    fail_start = 0

    # "FAIL" kelimesi geçen ilk sütunun indeksini bul
    for i in range(2, total_cols + 1):
        val = str(ws.cell(row=2, column=i).value).upper()
        if "FAIL" in val:
            fail_start = i
            break

    # 1. satıra grup başlıkları yerleştirme
    if fail_start > 2:
        # "Downlock Sensor" grubu için hücreleri birleştir
        ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=fail_start - 1)
        cell = ws.cell(row=1, column=2)
        cell.value = "NLG GEAR Downlock Sensor"
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = header_fill
        cell.border = thin_border

        # "Sensor Fail" grubu için hücreleri birleştir
        ws.merge_cells(start_row=1, start_column=fail_start, end_row=1, end_column=total_cols - 2)
        cell = ws.cell(row=1, column=fail_start)
        cell.value = "Sensor Fail"
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = header_fill
        cell.border = thin_border

        # Sağdaki iki sütun için boş grup başlığı hücreleri
        ws.cell(row=1, column=total_cols - 1).value = ""
        ws.cell(row=1, column=total_cols).value = ""
        ws.cell(row=1, column=total_cols - 1).border = thin_border
        ws.cell(row=1, column=total_cols).border = thin_border

    # Sol tarafa numara sütunu ekle
    ws.insert_cols(1)
    ws.cell(row=2, column=1).value = ""  # Başlık hücresi boş
    ws.column_dimensions["A"].width = 5

    # Her veri satırı için sıra numarası ekle
    for idx, row in enumerate(ws.iter_rows(min_row=3, min_col=2), start=0):
        cell = ws.cell(row=idx + 3, column=1)
        cell.value = idx
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # 2. satır başlık hücrelerini biçimlendir
    for cell in ws[2]:
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = header_fill
        cell.border = thin_border

    # 3. satırdan itibaren veri hücrelerini biçimlendir
    for row in ws.iter_rows(min_row=3, min_col=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            # Renk kodlaması
            if isinstance(cell.value, str):
                if cell.value.startswith("0="):
                    cell.fill = green_fill
                elif cell.value.startswith("1="):
                    cell.fill = red_fill

    # Sütun genişliği ve satır yüksekliği ayarla
    for i, col in enumerate(ws.columns, 1):
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(i)].width = min(max_length + 4, 30)

    for row in ws.iter_rows():
        ws.row_dimensions[row[0].row].height = 22  # Satır yüksekliği

    wb.save(filepath)

# Excel tablosunu oluşturma fonksiyonu
def create_table():
    try:
        sensor_count = int(entry_sensor_count.get())
        if sensor_count <= 0:
            raise ValueError
    except ValueError:
        messagebox.showerror("Error", "Enter a valid sensor count.")
        return

    # Kullanıcıdan girilen sensör adlarını al
    sensor_names = [entry.get().strip().upper() for entry in sensor_entries if entry.get().strip()]
    if len(sensor_names) != sensor_count:
        messagebox.showerror("Error", "Please enter all sensor names.")
        return

    # Her sensör için 0 ve 1 değerlerinin anlamlarını sor
    sensor_state_0_meanings = []
    sensor_state_1_meanings = []
    for name in sensor_names:
        s0 = simpledialog.askstring("Input", f"What does 0 mean for {name}?")
        if s0 is None: return
        s1 = simpledialog.askstring("Input", f"What does 1 mean for {name}?")
        if s1 is None: return
        sensor_state_0_meanings.append(s0.upper())
        sensor_state_1_meanings.append(s1.upper())

    # "FAIL" durumları için anlamlar
    fail_0 = simpledialog.askstring("Input", "What does 0 mean for FAIL?")
    if fail_0 is None: return
    fail_1 = simpledialog.askstring("Input", "What does 1 mean for FAIL?")
    if fail_1 is None: return

       # Sütun başlıklarını oluştur (ALT ALTA olacak şekilde)
    state_columns = [f"DOWNLOCK SENSOR {i+1}\nMCU{i+1}" for i in range(sensor_count)]
    fail_columns = [f"MCU{i+1}\nSENSOR FAIL{i+1}" for i in range(sensor_count)]
    all_columns = state_columns + fail_columns


    # Tüm kombinasyonları oluştur (0-1 bit dizileri)
    total_bits = sensor_count * 2
    combinations = list(itertools.product([0, 1], repeat=total_bits))
    data = []

    # Kombinasyonları metin olarak formatla (örn: "0=UNLOCKED")
    for combo in combinations:
        sensor_states = combo[::2]  # çift indeks: sensor
        fail_states = combo[1::2]   # tek indeks: fail

        formatted_states = [
            f"{v}={sensor_state_0_meanings[i] if v == 0 else sensor_state_1_meanings[i]}"
            for i, v in enumerate(sensor_states)
        ]
        formatted_fails = [f"{v}={fail_0 if v == 0 else fail_1}" for v in fail_states]
        data.append(formatted_states + formatted_fails)

    # Veri çerçevesini oluştur ve boş başlıklı 2 sütun ekle
    df = pd.DataFrame(data, columns=all_columns)
    df["DOWNLOK SENSOR STS"] = ""
    df["DOWNLOCK SENSOR FAIL STS"] = ""

    # Excel'e yaz ve biçimlendir
    output_file = "Sensor_Validation_Table.xlsx"
    df.to_excel(output_file, index=False)
    format_excel(output_file)
    messagebox.showinfo("Success", "Excel file created:\n" + os.path.abspath(output_file))

# --- GUI Arayüzü ---
root = tk.Tk()
root.title("Sensor Table Generator")
root.geometry("600x400")
root.resizable(False, False)

# Sensör sayısı girişi
tk.Label(root, text="Number of Sensors:").pack()
entry_sensor_count = tk.Entry(root)
entry_sensor_count.pack()

# Sensör adlarını dinamik olarak ekleyen fonksiyon
def update_sensor_entries():
    for widget in sensor_frame.winfo_children():
        widget.destroy()
    try:
        count = int(entry_sensor_count.get())
    except ValueError:
        return
    global sensor_entries
    sensor_entries = []
    for i in range(count):
        entry = tk.Entry(sensor_frame)
        entry.insert(0, f"Sensor{i+1}")
        entry.pack()
        sensor_entries.append(entry)

# Butonlar ve çerçeveler
tk.Button(root, text="Enter Sensor Names", command=update_sensor_entries).pack()
sensor_frame = tk.Frame(root)
sensor_frame.pack()
tk.Button(root, text="Generate Table and Save to Excel", command=create_table).pack()

# Uygulama döngüsü
root.mainloop()